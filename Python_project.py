from openpyxl.workbook import Workbook
import os
from openpyxl import load_workbook
import matplotlib.pyplot as pt
from os import listdir
from os import path
filepath="F:\Python-project\exelfoldersheets"
def filelist():
    for file in listdir(filepath):
        if path.isfile(path.join(filepath,file)):
             print(file)  

def fullprog():
    print("chose any 2 files to compare from existing file")
    filelist()
    x=input("enter the name of file 1=")
    y=input("enter the name of file 2=")
    wb1=load_workbook(str(x)+'.xlsx')
    wb2=load_workbook(str(y)+'.xlsx')
    ws1=wb1.active
    ws2=wb2.active

    item1=ws1['A']
    price1=ws1['B']
    quty1=ws1['C']

    item2=ws2['A']
    price2=ws2['B']
    quty2=ws2['C']

    itemlist1 = []
    itemlist2 = []
    pricelist1=[]
    pricelist2=[]
    qtylist1=[]
    qtylist2=[]
    profit1=[]
    profit2=[]
    diffofqty = []

    for i in item1:
        itemlist1.append(i.value)
    for i in item2:
        itemlist2.append(i.value)    
    for i in quty1:
        qtylist1.append(i.value)
    for i in quty2:
        qtylist2.append(i.value)
    for i in price1:
        pricelist1.append(i.value)
    for i in price2:
        pricelist2.append(i.value)   

    s1=max(qtylist1)
    s2=max(qtylist2)

    for i in range(0, len(qtylist1)):
        diffofqty.append(qtylist1[i] - qtylist2[i])
    x1=diffofqty    
    for i in range(0,len(qtylist1)):
        profit1.append(qtylist1[i]*pricelist1[i])
        z1=profit1
    for i in range(0,len(qtylist1)):
        profit2.append(qtylist2[i]*pricelist1[i])
        z2=profit2    

    total1=sum(z1)
    total2=sum(z2)

    max_index1 = qtylist1.index(max(qtylist1))
    max_index2 = qtylist2.index(max(qtylist2))

    m1=("A"+str(max_index1+1))
    m2=("A"+str(max_index2+1))
    print(f"total sales profit in sheet 1 is Rs.{total1}")
    print(f"total sales profit in sheet 2 is Rs.{total2}")
    print(f"phone sold in max qty in 1st sheet  was {(ws1[m1].value)} and {s1} units were sold ")
    print(f"phone sold in max qty in 2nd sheet was {(ws2[m2].value)} and {s2} units were sold")

    def graph1():
        pt.plot(itemlist1,x1)
        pt.title('rise and fall of sales')
        pt.plot()
        pt.show()
    def graph2():
        pt.bar(itemlist1,qtylist1)
        pt.title('sales in 1st sheet')
        pt.show()  

    def graph3():
        pt.bar(itemlist1,qtylist2)
        pt.title('sales in 2nd list')
        pt.show()    
    graph2()
    graph3()
    graph1()
    def ddd():
        ac1=('  sales in sheet 1='+str(total1))
        ac2=('  sales in sheet 1='+str(total2))
        zz='|--------|'
        with open('records.txt','a')as file:
            file.write(ac1)
            file.write(zz)
            file.write(ac2)
    ddd()        



def call():
    while True:
        print("enter any key to contine or 1 to exit")
        number=input()
        if number !='1':
            fullprog()
            zzz=input('enterany key to continue or 2 to deleat all data saved=')
            if zzz =='2':
                    os.remove('records.txt')
        else:
            break           

def login():
    pswrd="0987"
    input("enter the user name=")
    alenda=input("enter the password=")
    if alenda==pswrd:
        call()
    else:
        print("wrong password try again")
        login()    

login()        
